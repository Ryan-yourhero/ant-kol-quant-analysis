"""
每日 AI 分析报告服务
====================
从「成表」的 Excel（output/YYYYMMDD.xlsx）重建 TradeRecord 列表，
调用 src.parser.daily_report.analyze_daily 生成每日复盘报告。

支持：
- 列出所有有数据的日期 + 报告生成状态（历史）
- 异步批量生成报告（补跑历史日期）
- 读取某日期的报告内容
"""

from __future__ import annotations

import glob
import logging
import os
import re
import sys
import threading
from datetime import datetime
from typing import List, Optional

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from src.parser.models import TradeRecord  # noqa: E402

logger = logging.getLogger("backend.report_service")

OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# ---- 线程安全状态 ----
_lock = threading.Lock()
_state = {
    "status": "idle",  # idle / generating / success / failed
    "message": "",
    "total": 0,
    "done": 0,
    "current_date": None,
    "failed_dates": [],
    "started_at": None,
    "finished_at": None,
}


# ============================================================
#  从 Excel 重建 TradeRecord
# ============================================================

# 与 excel_exporter.EXCEL_COLUMNS 的表头顺序对齐
_EXCEL_HEADERS = [
    "大V昵称",
    "收益率",
    "发布时间",
    "动态正文",
    "操作类型",
    "操作状态",
    "基金名称",
    "买入金额（元）",
    "卖出份额（份）",
    "采集时间",
    "备注",
]


def _clean_amount(v) -> Optional[str]:
    """清洗金额：去掉千分位逗号，把 '--'/'-'/'None' 等占位符归为 None。"""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "--", "-", "None", "nan", "N/A"):
        return None
    return s.replace(",", "")


def load_records_from_excel(date_str: str) -> List[TradeRecord]:
    """从 output/YYYYMMDD.xlsx 重建 TradeRecord 列表（成表数据）。"""
    import openpyxl

    cleaned = date_str.replace("-", "")
    excel_path = os.path.join(OUTPUT_DIR, f"{cleaned}.xlsx")
    if not os.path.exists(excel_path):
        logger.warning("Excel 不存在，无法重建记录: %s", excel_path)
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    records: List[TradeRecord] = []
    header: Optional[list] = None

    try:
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = list(row)
                continue
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            data = dict(zip(header, row))
            operation_type = data.get("操作类型")
            remark = data.get("备注")
            fund_name = data.get("基金名称")

            # 转换操作：把 remark="转换" 映射到 转换前/后 基金名称，供方向汇总区分转入/转出
            convert_from = None
            convert_to = None
            if remark == "转换":
                if operation_type == "卖出":
                    convert_from = fund_name
                elif operation_type == "买入":
                    convert_to = fund_name

            records.append(
                TradeRecord(
                    kol_name=data.get("大V昵称"),
                    yield_rate=data.get("收益率"),
                    publish_time=data.get("发布时间"),
                    opinion_text=data.get("动态正文"),
                    operation_type=operation_type,
                    operation_status=data.get("操作状态"),
                    fund_name=fund_name,
                    buy_amount=_clean_amount(data.get("买入金额（元）")),
                    sell_shares=_clean_amount(data.get("卖出份额（份）")),
                    collect_time=data.get("采集时间"),
                    remark=remark,
                    convert_from_fund=convert_from,
                    convert_to_fund=convert_to,
                )
            )
    finally:
        wb.close()

    logger.info("从 Excel 重建 %d 条记录（日期 %s）", len(records), cleaned)
    return records


def _load_crawl_status(date_str: str) -> dict:
    """从 output/raw_pages_YYYYMMDD_*.json 读取最新一次的爬虫元数据，判断数据完整性。

    数据完整性依据爬虫元数据（stop_type / bottom_marker_detected / expand_remaining），
    而不是凭"某个历史大V今天没出现"去推测。
    """
    import json

    cleaned = date_str.replace("-", "")
    pattern = os.path.join(OUTPUT_DIR, f"raw_pages_{cleaned}_*.json")
    files = sorted(glob.glob(pattern))
    if not files:
        return {"available": False, "integrity": "unknown"}

    latest = files[-1]
    try:
        with open(latest, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        logger.warning("读取爬虫元数据失败: %s", e)
        return {"available": False, "integrity": "unknown"}

    stop_type = data.get("stop_type", "unknown")
    bottom_detected = bool(data.get("bottom_marker_detected", False))
    expand_remaining = data.get("expand_remaining_visible", 0)
    expand_failed = data.get("expand_permanently_failed", 0)

    complete = stop_type == "bottom" and bottom_detected and expand_remaining == 0
    integrity = "complete" if complete else "incomplete"

    return {
        "available": True,
        "integrity": integrity,
        "stop_type": stop_type,
        "bottom_detected": bottom_detected,
        "expand_remaining": expand_remaining,
        "expand_failed": expand_failed,
        "source_file": os.path.basename(latest),
    }


# ============================================================
#  历史列表
# ============================================================

def _dates_with_data() -> List[str]:
    """扫描 output/????????.xlsx 得到所有有数据的日期（YYYYMMDD 降序，最近日期在前）。"""
    dates = set()
    for p in glob.glob(os.path.join(OUTPUT_DIR, "????????.xlsx")):
        m = re.match(r"^(\d{8})\.xlsx$", os.path.basename(p))
        if m:
            dates.add(m.group(1))
    return sorted(dates, reverse=True)


def _report_path(date_str: str) -> str:
    cleaned = date_str.replace("-", "")
    return os.path.join(OUTPUT_DIR, f"daily_report_{cleaned}.md")


def _excel_record_count(date_str: str) -> int:
    import openpyxl

    excel_path = os.path.join(OUTPUT_DIR, f"{date_str.replace('-', '')}.xlsx")
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        n = wb.active.max_row - 1
        wb.close()
        return max(0, n)
    except Exception as e:
        logger.warning("读取 Excel 行数失败: %s", e)
        return 0


def list_report_history() -> dict:
    """列出所有有数据的日期及其报告生成状态。"""
    items = []
    for d in _dates_with_data():
        rp = _report_path(d)
        has_report = os.path.exists(rp)
        items.append(
            {
                "date": f"{d[:4]}-{d[4:6]}-{d[6:]}",
                "record_count": _excel_record_count(d),
                "has_report": has_report,
                "report_path": rp if has_report else None,
            }
        )
    with _lock:
        status = dict(_state)
    return {"items": items, "status": status}


def get_report_content(date_str: str) -> Optional[str]:
    """读取某日期的报告内容（Markdown 原文）。"""
    rp = _report_path(date_str)
    if not os.path.exists(rp):
        return None
    with open(rp, "r", encoding="utf-8") as f:
        return f.read()


# ============================================================
#  异步生成
# ============================================================

def _generate_one(date_str: str):
    """生成单个日期的报告，返回 (ok, info)。"""
    records = load_records_from_excel(date_str)
    if not records:
        return False, "该日期无记录"

    from src.parser import analyze_daily
    from .historical_context_service import build as build_historical_context

    historical_context = None
    try:
        historical_context = build_historical_context(date_str, records)
        logger.info("[REPORT] %s 历史上下文构建完成: %d 大V, %d 方向",
                    date_str, len(historical_context.get("kols", [])), len(historical_context.get("directions", [])))
    except Exception as e:
        logger.warning("[REPORT] %s 历史上下文构建失败（降级为无历史对比）: %s", date_str, e)

    # 数据完整性（爬虫元数据），独立于历史聚合，总是注入
    crawl_status = _load_crawl_status(date_str)
    if historical_context is None:
        historical_context = {}
    historical_context["crawl_status"] = crawl_status

    path = analyze_daily(
        records,
        output_dir=OUTPUT_DIR,
        date_str=date_str,
        historical_context=historical_context,
    )
    if path:
        return True, path
    return False, "LLM 未配置或调用失败"


def _run(dates: List[str]):
    now = datetime.now().isoformat()
    with _lock:
        _state.update(
            status="generating",
            message=f"正在生成 {len(dates)} 个日期...",
            total=len(dates),
            done=0,
            current_date=None,
            failed_dates=[],
            started_at=now,
            finished_at=None,
        )

    for d in dates:
        with _lock:
            _state["current_date"] = d
        logger.info("[REPORT] 开始生成 %s", d)
        try:
            ok, info = _generate_one(d)
        except Exception as e:
            logger.error("[REPORT] %s 生成异常: %s", d, e, exc_info=True)
            ok, info = False, str(e)
        with _lock:
            if ok:
                _state["done"] += 1
            else:
                _state["failed_dates"].append({"date": d, "error": info})
            _state["current_date"] = None

    with _lock:
        failed = len(_state["failed_dates"])
        _state["status"] = "success" if failed == 0 else "failed"
        _state["message"] = f"完成 {_state['done']}/{_state['total']}，失败 {failed}"
        _state["finished_at"] = datetime.now().isoformat()


def generate_reports_async(dates: Optional[List[str]] = None) -> dict:
    """启动后台批量生成。dates=None 表示全部有数据的日期。"""
    with _lock:
        if _state["status"] == "generating":
            return {"ok": False, "message": "正在生成中，请稍候"}

    if not dates:
        target = _dates_with_data()
    else:
        target = [d.replace("-", "") for d in dates if d]

    if not target:
        return {"ok": False, "message": "没有可生成的日期"}

    thread = threading.Thread(target=_run, args=(target,), daemon=True)
    thread.start()
    return {"ok": True, "message": f"已启动 {len(target)} 个日期的报告生成"}
