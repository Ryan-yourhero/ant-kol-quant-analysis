"""Test full parse_md_to_excel flow"""
import sys
sys.path.insert(0, ".")

from src.parser import parse_md_to_excel

md_path = "output/screen_dump_20260805_162701.md"
print(f"Testing: {md_path}")
print()

excel_path, result = parse_md_to_excel(md_path, use_ai=False)

print(f"\nResult:")
print(f"  Excel path: {excel_path}")
print(f"  Total records: {result.total_records}")
print(f"  AI used: {result.ai_used}")
print()

if excel_path:
    print(f"Excel file created: {excel_path}")
    # Verify by reading it back
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    print(f"  Sheet name: {ws.title}")
    print(f"  Rows: {ws.max_row}")
    print(f"  Cols: {ws.max_column}")
    print(f"  Freeze pane: {ws.freeze_panes}")
    print(f"  Auto filter: {ws.auto_filter.ref}")
    print()
    # Print header
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"  Headers: {headers}")
    print()
    # Print first 3 rows
    for row in range(2, min(5, ws.max_row + 1)):
        vals = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
        print(f"  Row {row}: {vals}")
