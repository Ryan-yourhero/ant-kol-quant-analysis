"""Print Excel summary"""
from openpyxl import load_workbook
wb = load_workbook("data/excel/20260811.xlsx")
ws = wb.active
print(f"Sheet: {ws.title}  |  Rows: {ws.max_row} (1 header + {ws.max_row-1} data)")
print(f"Freeze: {ws.freeze_panes}  |  Filter: {ws.auto_filter.ref}")
print()
for r in range(2, ws.max_row+1):
    kol = ws.cell(r, 1).value or ""
    op = ws.cell(r, 6).value or ""
    fund = ws.cell(r, 8).value or ""
    buy = ws.cell(r, 9).value or ""
    sell = ws.cell(r, 10).value or ""
    cfrom = ws.cell(r, 11).value or ""
    cto = ws.cell(r, 12).value or ""
    ti = ws.cell(r, 4).value or ""
    today = ws.cell(r, 18).value or ""

    parts = [kol, op]
    if fund: parts.append(fund)
    if buy: parts.append(f"buy={buy}")
    if sell: parts.append(f"sell={sell}")
    if cfrom and op == "转换":
        parts.append(f"{cfrom} -> {cto}")
    parts.append(ti)

    print(f"  {r-1:2d}. {' | '.join(parts)}")
