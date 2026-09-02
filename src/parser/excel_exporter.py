"""
Excel 导出器 — 将解析结果导出为 .xlsx
"""

from __future__ import annotations

import os
import re
import datetime
import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import TradeRecord, ParseResult
from .rule_parser import extract_segments_from_md
from .ai_parser import parse_full_md
from .deduplicator import deduplicate

logger = logging.getLogger("parser.excel_exporter")

# ============================================================
#  列定义
# ============================================================

EXCEL_COLUMNS = [
    ("大V昵称", "kol_name", 16),
    ("收益率", "yield_rate", 14),
    ("发布时间", "publish_time", 12),
    ("动态正文", "opinion_text", 50),
    ("操作类型", "operation_type", 10),
    ("操作状态", "operation_status", 10),
    ("基金名称", "fund_name", 28),
    ("买入金额（元）", "buy_amount", 14),
    ("卖出份额（份）", "sell_shares", 14),
    ("采集时间", "collect_time", 20),
    ("备注", "remark", 16),
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=False)
OPINION_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# ============================================================
#  核心导出
# ============================================================

def _record_to_row(record: TradeRecord) -> List[Optional[str]]:
    d = record.model_dump()
    return [d.get(key) for _, key, _ in EXCEL_COLUMNS]


def export_to_excel(
    records: List[TradeRecord],
    output_path: str,
) -> str:
    """将 TradeRecord 列表导出为 .xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = "大V每日操作"

    # ---- 写表头 ----
    headers = [col[0] for col in EXCEL_COLUMNS]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # ---- 找 opinion_text 列号 ----
    opinion_col_idx = None
    for col_idx, (_, key, _) in enumerate(EXCEL_COLUMNS, 1):
        if key == "opinion_text":
            opinion_col_idx = col_idx
            break

    # ---- 不在这里排序：排序由调用方按 MD 中大V出现顺序处理 ----

    # ---- 写数据 ----
    for row_idx, record in enumerate(records, 2):
        row_data = _record_to_row(record)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col_idx == opinion_col_idx:
                cell.alignment = OPINION_ALIGNMENT
            else:
                cell.alignment = DATA_ALIGNMENT

    # ---- 自动列宽 ----
    for col_idx, (_, _, default_width) in enumerate(EXCEL_COLUMNS, 1):
        max_width = default_width
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                char_width = sum(2 if ord(c) > 127 else 1 for c in str(val))
                max_width = max(max_width, char_width + 2)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_width, 60)

    # ---- 冻结第一行 ----
    ws.freeze_panes = "A2"

    # ---- 自动筛选 ----
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{len(records) + 1}"

    # ---- 保存 ----
    output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        # 文件被占用，加时间戳另存
        base, ext = os.path.splitext(output_path)
        ts = datetime.datetime.now().strftime("%H%M%S")
        output_path = f"{base}_{ts}{ext}"
        wb.save(output_path)
        logger.warning("原文件被占用，另存为: %s", output_path)

    logger.info(f"Excel 已保存: {output_path} ({len(records)} 行)")
    return output_path


# ============================================================
#  主流程：MD → Excel
# ============================================================

def parse_md_to_excel(
    md_path: str,
    *,
    excel_dir: Optional[str] = None,
    use_ai: Optional[bool] = None,
) -> Tuple[Optional[str], ParseResult]:
    """
    一站式：screen_dump MD → Excel

    流程：
      读取 MD
      → AI 完整语义解析（返回嵌套 posts JSON）
      → Python 展平为 TradeRecord 列表
      → Pydantic 校验 + 去重 + 补充采集时间
      → 导出 Excel

    Args:
        md_path: screen_dump_*.md 文件路径
        excel_dir: Excel 输出目录（默认 output/）
        use_ai: 是否使用 AI（None=自动检测）

    Returns:
        (excel_path, ParseResult)
    """
    # ---- Step 0: 读取 MD ----
    if not os.path.exists(md_path):
        result = ParseResult(
            source_md=md_path,
            parse_time=datetime.datetime.now().isoformat(timespec="seconds"),
        )
        logger.error("MD 文件不存在: %s", md_path)
        return None, result

    with open(md_path, "r", encoding="utf-8") as f:
        md_text = f.read()

    # ---- Step 1: AI 端到端解析 ----
    should_use_ai = use_ai
    if should_use_ai is None:
        from .llm_client import is_configured
        should_use_ai = is_configured()

    if not should_use_ai:
        # 无 AI：退回纯规则解析
        logger.info("AI 未配置，使用纯规则解析")
        from .ai_parser import _dicts_to_records
        raw_records = extract_segments_from_md(md_text, md_path=md_path)
        raw_records = deduplicate(raw_records)
        records = _dicts_to_records(raw_records)
        ai_used = False
    else:
        # AI 端到端解析
        logger.info("Step 1/3: AI 端到端解析完整 MD ...")
        from .ai_parser import parse_full_md
        try:
            records = parse_full_md(md_text)
            ai_used = True
        except Exception as e:
            logger.warning("AI 解析异常（退回纯规则）: %s", e)
            from .ai_parser import _dicts_to_records
            raw_records = extract_segments_from_md(md_text, md_path=md_path)
            raw_records = deduplicate(raw_records)
            records = _dicts_to_records(raw_records)
            ai_used = False

    # ---- Step 2: Python 校验 & 去重 ----
    logger.info("Step 2/3: Python 校验 & 去重（%d 条记录）...", len(records))
    records = _validate_and_dedup(records)
    logger.info("校验后: %d 条", len(records))

    # ---- Step 3: 导出 Excel ----
    logger.info("Step 3/3: 导出 Excel（%d 条记录）...", len(records))

    if excel_dir is None:
        _parser_dir = os.path.dirname(os.path.abspath(__file__))
        _project_root = os.path.dirname(os.path.dirname(_parser_dir))
        excel_dir = os.path.join(_project_root, "output")

    date_match = re.search(r"(\d{8})", os.path.basename(md_path))
    date_str = date_match.group(1) if date_match else datetime.date.today().strftime("%Y%m%d")
    excel_name = f"{date_str}.xlsx"
    excel_path = os.path.join(excel_dir, excel_name)

    # 按大V在 MD 中出现顺序排序
    _sort_by_kol_order(records, md_text)

    logger.info("Step 3/3: 导出 Excel → %s", excel_path)
    try:
        saved_path = export_to_excel(records, excel_path)
    except Exception as e:
        logger.error("Excel 导出失败: %s", e)
        result = ParseResult(records=records, source_md=md_path, ai_used=ai_used)
        return None, result

    # ---- Step 4: 写入 MySQL ----
    _try_mysql_save(records, md_text, md_path)

    result = ParseResult(records=records, source_md=md_path, ai_used=ai_used)
    return saved_path, result


def _validate_and_dedup(records: List[TradeRecord]) -> List[TradeRecord]:
    """Python 侧校验 + 去重 + 补充采集时间"""
    from .deduplicator import deduplicate

    # Pydantic 已在 TradeRecord 构造时校验。这里做业务校验。
    collect_time = datetime.datetime.now().isoformat(timespec="seconds")

    validated: List[TradeRecord] = []
    for r in records:
        # 补充采集时间（AI 不输出此字段）
        r.collect_time = collect_time

        # operation_type 枚举校验
        if r.operation_type and r.operation_type not in ("买入", "卖出", "撤销", "定投"):
            logger.warning("无效操作类型: %s，跳过", r.operation_type)
            continue

        # 过滤空记录：没有基金名且没有金额的跳过
        if not r.fund_name and not r.buy_amount and not r.sell_shares:
            logger.warning("空记录（无基金名/无金额），跳过")
            continue

        # buy_amount / sell_shares 格式化
        if r.buy_amount:
            try:
                r.buy_amount = f"{float(r.buy_amount):.2f}"
            except (ValueError, TypeError):
                pass
        if r.sell_shares:
            try:
                r.sell_shares = f"{float(r.sell_shares):.2f}"
            except (ValueError, TypeError):
                pass

        validated.append(r)

    # 跨记录去重（同大V+同操作+同基金+同金额 → 只保留第一条）
    seen = set()
    deduped: List[TradeRecord] = []
    for r in validated:
        key = (r.kol_name, r.operation_type, r.fund_name, r.buy_amount, r.sell_shares)
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    return deduped


def _extract_kol_order(md_text: str) -> dict:
    """从 MD 中提取大V出现顺序，返回 {kol_name: position}"""
    import re as _re

    _blacklist = {
        "关注", "发现", "讨论区", "热议话题", "学理财", "资讯", "同路人",
        "真实财有趣", "我的关注", "最新", "热门", "全部", "推荐",
        "今日操作", "原创", "转发", "分享", "收藏", "评论", "点赞",
        "求解读", "回复", "催一下", "查看详情", "立即查看",
        "加载中", "暂无数据", "暂无更多内容", "展开", "理财盘友圈", "记一下",
        "首页", "理财", "资产", "消息", "我的", "返回", "搜索", "设置",
        "更多", "关闭", "取消", "确定",
    }

    order: dict = {}
    pos = 0
    lines = md_text.split("\n")
    for i, line in enumerate(lines):
        stripped = line.strip()
        if (
            len(stripped) >= 2
            and len(stripped) <= 12
            and _re.match(r"^[\u4e00-\u9fa5A-Za-z0-9_]+$", stripped)
            and stripped not in _blacklist
            and not _re.fullmatch(r"\d+", stripped)
            and "展开今日" not in stripped
            and "人求解读" not in stripped
        ):
            has_yield = any(
                "%" in lines[j] for j in range(i + 1, min(i + 6, len(lines)))
            )
            if has_yield:
                if stripped not in order:
                    order[stripped] = pos
                    pos += 1
    return order


def _sort_by_kol_order(records, md_text: str) -> None:
    """按大V在 MD 中出现顺序原地排序 records"""
    order = _extract_kol_order(md_text)
    if not order:
        return

    def _key(rec):
        name = rec.kol_name or ""
        # 处理 (待定) 后缀：用基础名查顺序，(待定) 同名内排后面
        base = name.replace("(待定)", "")
        pending = "(待定)" in name
        pos = order.get(base, 9999)
        return (pos, pending, name)

    records.sort(key=_key)


def _try_mysql_save(
    records: List[TradeRecord],
    md_text: str,
    md_path: str,
) -> None:
    """尝试写入 MySQL。失败不影响流程，只记日志。"""
    try:
        from ..storage.db_storage import is_configured as mysql_configured
        from ..storage.db_storage import save_records as mysql_save
        from ..storage.db_storage import init_db
        from ..storage.db_storage import ensure_mysql_running

        if not mysql_configured():
            logger.debug("MySQL 未配置，跳过")
            return

        if not ensure_mysql_running():
            logger.warning("MySQL 无法启动/连接，跳过写库（不影响 Excel）")
            return

        init_db()
        logger.info("Step 4/4: 写入 MySQL ...")
        mysql_save(records, md_text, md_path)
    except Exception as e:
        logger.warning("MySQL 写入异常（不影响 Excel）: %s", e)
