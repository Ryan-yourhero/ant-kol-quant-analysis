"""
蚂蚁财富大V交易数据 Excel 持久化存储模块。

规则：
  1. 每次采集完成后，**保留**原 JSON 文件不变；
  2. 将交易记录追加写入 output/ant_finance_trades.xlsx（2 个 Sheet）；
  3. Excel 不存在自动创建；已存在则追加写入，不覆盖历史；
  4. 按 (大V名称 + 操作时间 + 操作类型 + 基金名称 + 金额) 去重；
  5. 即使本次解析交易数为 0，也会在 raw_data 中保留一整份 JSON 内容，
     方便后续排查解析问题。

Sheet 结构：
  Sheet1: trade_records
    id / 采集日期 / 数据来源 / 大V名称 / 操作时间 / 操作类型 /
    基金名称 / 源基金 / 目标基金 / 操作金额 / 金额单位 /
    观点文本 / 点赞数 / 评论数 / 置信度 / 原始操作文本 / 采集时间

  Sheet2: raw_data
    采集时间 / JSON文件名 / JSON原始内容

使用：
  1) 在 main.py 中调用：
      from core.excel_storage import persist_json_to_excel
      persist_json_to_excel(json_path)

  2) 独立 CLI：
      python core/excel_storage.py output/kol_trades_20260804_235016.json
"""
from __future__ import annotations

import os
import sys
import json
import argparse
import datetime
import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


logger = logging.getLogger("excel_storage")


# ============================================================
#  Sheet 列契约（严格固定顺序、严格列头文字；空白列也要保留）
# ============================================================

TRADE_HEADERS: List[str] = [
    "id",              # 1  自增，每次新行 = max(现有id)+1
    "采集日期",         # 2  YYYY-MM-DD
    "数据来源",         # 3  默认 "蚂蚁财富"
    "大V名称",         # 4
    "操作时间",         # 5  原始 timestamp，不填则保持空（与去重对称）
    "操作类型",         # 6  BUY / SELL / TRANSFER / CANCEL
    "基金名称",         # 7  BUY/SELL 非空；TRANSFER 空；CANCEL 存 "撤销类型|基金名"
    "源基金",           # 8  TRANSFER 非空
    "目标基金",         # 9  TRANSFER 非空
    "操作金额",         # 10 纯数值；TRANSFER 用 "源金额 / 目标金额"
    "金额单位",         # 11 元 / 份 / 份/元
    "观点文本",         # 12 从 operation_text 中抽取的"可能是观点"的长文（>30字才写）
    "点赞数",           # 13 预留空
    "评论数",           # 14 预留空
    "置信度",           # 15 0~1
    "原始操作文本",     # 16 完整 operation_text（不可丢失，AI 可复核）
    "采集时间",         # 17 YYYY-MM-DD HH:MM:SS
]

RAW_HEADERS: List[str] = [
    "采集时间",       # 1
    "JSON文件名",     # 2
    "JSON原始内容",   # 3
]

DEFAULT_DATA_SOURCE = "蚂蚁财富"
DEFAULT_XLSX_BASENAME = "ant_finance_trades.xlsx"


# ============================================================
#  通用小工具
# ============================================================

def _s(v: Any, default: str = "") -> str:
    if v is None:
        return default
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return f"{v:.6f}".rstrip("0").rstrip(".")
    return str(v)


def _f(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    import re
    m = re.search(r"([\d,]+(?:\.\d+)?)", str(v).replace("，", ","))
    try:
        return float(m.group(1).replace(",", "")) if m else 0.0
    except Exception:
        return 0.0


def _split_datetime(iso_or_dt: str) -> Tuple[str, str]:
    s = (iso_or_dt or "").strip().replace("T", " ")
    try:
        if len(s) >= 19:
            dt = datetime.datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
        elif len(s) >= 10:
            dt = datetime.datetime.strptime(s[:10], "%Y-%m-%d")
        else:
            raise ValueError
    except Exception:
        dt = datetime.datetime.now()
    return dt.strftime("%Y-%m-%d"), dt.strftime("%Y-%m-%d %H:%M:%S")


def _read_json(json_path: str) -> Tuple[Dict[str, Any], str, str]:
    """读取 JSON → (payload, 采集时间 str, JSON 原始文本)"""
    # utf-8-sig 可自动处理 PowerShell/某些编辑器写入的 BOM
    with open(json_path, "r", encoding="utf-8-sig") as f:
        raw_text = f.read()
    payload = json.loads(raw_text)
    if not isinstance(payload, dict):
        payload = {
            "records": payload if isinstance(payload, list) else [],
            "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        }
    generated_at = _s(
        payload.get("generated_at") or payload.get("collected_at") or ""
    ) or datetime.datetime.now().isoformat(timespec="seconds")
    return payload, generated_at, raw_text


def _extract_records(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    """兼容两种格式：records / collected_trades / operations"""
    for key in ("records", "collected_trades", "operations"):
        val = payload.get(key)
        if isinstance(val, list):
            return [r for r in val if isinstance(r, dict)]
    return []


# ============================================================
#  去重键（写入 ↔ 读取 必须完全对称）
# ============================================================

def _record_fund_for_dedup(rec: Dict[str, Any]) -> str:
    """根据操作类型决定去重用的"基金"字符串"""
    at = _s(rec.get("action_type") or rec.get("operation")).upper()
    if at == "TRANSFER":
        return f"{_s(rec.get('source_fund'))}|{_s(rec.get('target_fund'))}"
    if at == "CANCEL":
        return f"{_s(rec.get('cancel_type') or 'UNKNOWN')}|{_s(rec.get('fund') or rec.get('fund_name'))}"
    # BUY / SELL / 未知
    return _s(rec.get("fund") or rec.get("fund_name"))


def _record_amount_for_dedup(rec: Dict[str, Any]) -> str:
    """去重用纯数值金额（TRANSFER 为 "源/目标"）"""
    at = _s(rec.get("action_type") or rec.get("operation")).upper()
    if at == "TRANSFER":
        src = _f(rec.get("source_amount_value") or rec.get("source_amount"))
        tgt = _f(rec.get("target_amount_value") or rec.get("target_amount"))
        if src and tgt:
            return f"{_s(src)} / {_s(tgt)}"
        return _s(src) or _s(tgt)
    # 单金额
    av = rec.get("amount_value")
    if av not in (None, "", 0, 0.0):
        return _s(av)
    return _s(_f(rec.get("amount"))) if _f(rec.get("amount")) else ""


def _dedup_key(rec: Dict[str, Any]) -> Tuple[str, str, str, str, str]:
    """
    5 元唯一键：(大V名称, 操作时间, 操作类型, 基金, 金额)
    注意：操作时间**保持与记录一致（空就是空）**，不能用"采集时间"回填，否则去重失效。
    """
    kol = _s(rec.get("kol_name")).strip()
    ts = _s(rec.get("timestamp")).strip()
    at = _s(rec.get("action_type") or rec.get("operation")).strip().upper()
    fund = _record_fund_for_dedup(rec).strip()
    amt = _record_amount_for_dedup(rec).strip()
    return (kol, ts, at, fund, amt)


# ============================================================
#  行构造
# ============================================================

def _build_trade_row(
    next_id: int,
    rec: Dict[str, Any],
    collect_date: str,
    collect_time: str,
    data_source: str,
) -> List[Any]:
    """按 TRADE_HEADERS 顺序构造一行；严格 17 列，空白保持空字符串。"""
    kol = _s(rec.get("kol_name")).strip()
    ts = _s(rec.get("timestamp")).strip()  # 不空才写
    at = _s(rec.get("action_type") or rec.get("operation")).strip().upper()

    # --- 基金列 ---
    source_fund = _s(rec.get("source_fund"))
    target_fund = _s(rec.get("target_fund"))
    if at == "TRANSFER":
        fund_name = ""
    elif at == "CANCEL":
        fund_name = f"{_s(rec.get('cancel_type') or 'UNKNOWN')}|{_s(rec.get('fund') or rec.get('fund_name'))}"
    else:
        fund_name = _s(rec.get("fund") or rec.get("fund_name"))

    # --- 操作金额 + 单位 ---
    if at == "TRANSFER":
        src_val = _f(rec.get("source_amount_value") or rec.get("source_amount"))
        tgt_val = _f(rec.get("target_amount_value") or rec.get("target_amount"))
        if src_val and tgt_val:
            amount_str = f"{_s(src_val)} / {_s(tgt_val)}"
        else:
            amount_str = _s(src_val) or _s(tgt_val) or ""
        unit_str = f"{_s(rec.get('source_amount_unit') or '份')} / {_s(rec.get('target_amount_unit') or '元')}"
    else:
        av = rec.get("amount_value")
        if av not in (None, "", 0, 0.0):
            amount_str = _s(av)
        else:
            amount_str = _s(_f(rec.get("amount"))) if _f(rec.get("amount")) else ""
        unit_str = _s(rec.get("unit") or rec.get("amount_unit") or (
            "份" if at == "SELL" else "元" if at in ("BUY", "CANCEL") else ""
        ))

    # --- 观点文本 vs 原始操作文本 拆分 ---
    op_text = _s(rec.get("operation_text"))
    # 观点文本：从 operation_text 中抽取 ≥30 字的行，或整段超过 80 字就取它
    opinion_lines: List[str] = []
    for line in op_text.splitlines():
        if len(line) >= 30:
            opinion_lines.append(line)
    if len(opinion_lines) == 0 and len(op_text) >= 80:
        opinion_text = op_text[:800]
    else:
        opinion_text = "\n".join(opinion_lines)[:800]
    if len(opinion_text) == 0:
        opinion_text = ""  # 保持空

    # 原始操作文本：完整保留（超过 32k 做截断，Excel 单元格上限）
    raw_op_text = op_text
    if len(raw_op_text) > 32000:
        raw_op_text = raw_op_text[: 32000 - 16] + "…[TRUNCATED]"

    like_cnt = ""
    comment_cnt = ""
    confidence = _f(rec.get("confidence"))

    return [
        next_id,          # id
        collect_date,     # 采集日期
        data_source,      # 数据来源
        kol,              # 大V名称
        ts,               # 操作时间
        at,               # 操作类型
        fund_name,        # 基金名称
        source_fund,      # 源基金
        target_fund,      # 目标基金
        amount_str,       # 操作金额
        unit_str,         # 金额单位
        opinion_text,     # 观点文本
        like_cnt,         # 点赞数
        comment_cnt,      # 评论数
        confidence,       # 置信度
        raw_op_text,      # 原始操作文本
        collect_time,     # 采集时间
    ]


# ============================================================
#  ExcelStorage
# ============================================================

class ExcelStorage:
    def __init__(self, xlsx_path: str):
        self.xlsx_path = os.path.abspath(xlsx_path)
        os.makedirs(os.path.dirname(self.xlsx_path) or ".", exist_ok=True)

    # ----------------------------------------------------------
    def _open_or_create(self) -> Workbook:
        if os.path.exists(self.xlsx_path):
            wb = load_workbook(self.xlsx_path)
            for sheet_name, headers in (("trade_records", TRADE_HEADERS), ("raw_data", RAW_HEADERS)):
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=sheet_name)
                    self._write_header(ws, headers)
            return wb
        else:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "trade_records"
            self._write_header(ws1, TRADE_HEADERS)
            ws2 = wb.create_sheet("raw_data")
            self._write_header(ws2, RAW_HEADERS)
            return wb

    @staticmethod
    def _write_header(ws, headers: List[str]) -> None:
        has_content = any(ws.cell(row=1, column=i).value is not None for i in range(1, len(headers) + 1))
        if has_content:
            return
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        bold = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = bold
            c.fill = fill
            c.alignment = center
        ws.freeze_panes = "A2"

    # ----------------------------------------------------------
    @staticmethod
    def _current_max_id(ws_trade) -> int:
        if ws_trade.max_row <= 1:
            return 0
        best = 0
        for r in range(2, ws_trade.max_row + 1):
            v = ws_trade.cell(row=r, column=1).value
            try:
                vv = int(v) if v is not None else 0
            except (ValueError, TypeError):
                vv = 0
            if vv > best:
                best = vv
        return best

    # ----------------------------------------------------------
    def _load_trade_keys(self, ws_trade) -> set:
        keys: set = set()
        if ws_trade.max_row <= 1:
            return keys
        header = [ws_trade.cell(row=1, column=i).value for i in range(1, len(TRADE_HEADERS) + 1)]
        # 按列头名称索引（与 TRADE_HEADERS 顺序严格一致时等价）
        pos = {str(h): i for i, h in enumerate(header, 1) if h is not None}

        def col(name: str) -> Optional[int]:
            return pos.get(name)

        for r in range(2, ws_trade.max_row + 1):
            def gv(name):
                c = col(name)
                return _s(ws_trade.cell(row=r, column=c).value) if c else ""

            kol = gv("大V名称").strip()
            ts = gv("操作时间").strip()
            at = gv("操作类型").strip().upper()
            amt_cell = _s(gv("操作金额"))
            # 去重 "基金名称" 与写入对称：TRANSFER = 源|目标；CANCEL = 撤类型|基金；其他 = 基金名称
            if at == "TRANSFER":
                fund = f"{gv('源基金').strip()}|{gv('目标基金').strip()}"
            else:
                fund = gv("基金名称").strip()
            keys.add((kol, ts, at, fund, amt_cell.strip()))
        return keys

    # ----------------------------------------------------------
    def _append_raw(self, ws_raw, collect_time: str, json_path: str, raw_json_text: str) -> None:
        # raw_data 即使本次解析 0 条，也写一行。
        fname = os.path.basename(json_path)
        payload_text = raw_json_text
        if len(payload_text) > 32000:
            payload_text = payload_text[: 32000 - 16] + "…[TRUNCATED]"
        ws_raw.append([collect_time, fname, payload_text])

    # ----------------------------------------------------------
    #  主入口：写一次采集（= 一个 JSON 文件）
    # ----------------------------------------------------------
    def persist_json(
        self,
        json_path: str,
        data_source: str = DEFAULT_DATA_SOURCE,
    ) -> Dict[str, int]:
        if not os.path.exists(json_path):
            raise FileNotFoundError(json_path)

        payload, generated_at, raw_json_text = _read_json(json_path)
        records = _extract_records(payload)
        collect_date, collect_time = _split_datetime(generated_at)

        wb = self._open_or_create()
        ws_trade = wb["trade_records"]
        ws_raw = wb["raw_data"]

        existing_keys = self._load_trade_keys(ws_trade)
        next_id = self._current_max_id(ws_trade)

        trade_added = 0
        for rec in records:
            key = _dedup_key(rec)
            if key in existing_keys:
                continue
            existing_keys.add(key)
            next_id += 1
            row = _build_trade_row(next_id, rec, collect_date, collect_time, data_source)
            ws_trade.append(row)
            trade_added += 1

        # raw_data：每次采集必写一行（即使 0 条交易）
        self._append_raw(ws_raw, collect_time, json_path, raw_json_text)

        self._autosize(wb)
        wb.save(self.xlsx_path)

        stats = {
            "trade_records_added": trade_added,
            "trade_records_total": ws_trade.max_row - 1,
            "raw_data_rows_added": 1,
            "raw_data_rows_total": ws_raw.max_row - 1,
        }
        logger.info(
            f"持久化完成 → {self.xlsx_path}："
            f"trade +{stats['trade_records_added']}（累计 {stats['trade_records_total']}），"
            f"raw_data 累计 {stats['raw_data_rows_total']}，"
            f"本次 JSON 解析 {len(records)} 条记录。"
        )
        return stats

    # ----------------------------------------------------------
    def _autosize(self, wb: Workbook) -> None:
        trade_widths = [6, 12, 10, 16, 20, 10, 36, 36, 36, 16, 10, 60, 8, 8, 8, 60, 20]
        raw_widths = [20, 42, 80]
        for sheet_name, widths in (("trade_records", trade_widths), ("raw_data", raw_widths)):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            # 自动换行：观点文本 / 原始操作文本 / JSON原始内容
            wrap = Alignment(wrap_text=True, vertical="top")
            if sheet_name == "trade_records":
                for col_letter, col_idx in [("L", 12), ("P", 16)]:
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=col_idx).alignment = wrap
                conf_col = TRADE_HEADERS.index("置信度") + 1
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=conf_col).number_format = "0.000"
            if sheet_name == "raw_data":
                json_col = RAW_HEADERS.index("JSON原始内容") + 1
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=json_col).alignment = wrap


# ============================================================
#  便捷函数（main.py 直接调用）
# ============================================================

def persist_json_to_excel(
    json_path: str,
    xlsx_path: Optional[str] = None,
    data_source: str = DEFAULT_DATA_SOURCE,
) -> Dict[str, int]:
    """
    [兼容旧版] 单次采集结束后调用：对 records JSON 追加 trade_records + raw_data。
    新架构 v2 请使用 persist_raw_text_json_to_excel（只写 raw_data，不判断买卖）。
    """
    if not xlsx_path:
        candidate = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output", DEFAULT_XLSX_BASENAME)
        xlsx_path = candidate
    storage = ExcelStorage(xlsx_path)
    return storage.persist_json(json_path, data_source=data_source)


def persist_raw_text_json_to_excel(
    json_path: str,
    xlsx_path: Optional[str] = None,
    data_source: str = DEFAULT_DATA_SOURCE,
) -> Dict[str, int]:
    """
    v2 架构便捷入口：
      - 只写 raw_data（每次采集必写一行，保存原始 JSON 全文，便于后续 AIParser 复核）
      - 不解析 JSON → 不追加 trade_records（AI Parser 阶段再写）
    返回统计。
    """
    if not os.path.exists(json_path):
        raise FileNotFoundError(json_path)

    if not xlsx_path:
        candidate = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output", DEFAULT_XLSX_BASENAME)
        xlsx_path = candidate

    # 读取 JSON，拿 generated_at 作为采集时间
    try:
        with open(json_path, "r", encoding="utf-8-sig") as f:
            raw_text = f.read()
        payload = json.loads(raw_text)
        if isinstance(payload, dict):
            generated_at = _s(
                payload.get("generated_at") or payload.get("collected_at") or ""
            ) or datetime.datetime.now().isoformat(timespec="seconds")
        else:
            generated_at = datetime.datetime.now().isoformat(timespec="seconds")
    except Exception:
        raw_text = ""
        generated_at = datetime.datetime.now().isoformat(timespec="seconds")

    _, collect_time = _split_datetime(generated_at)

    storage = ExcelStorage(xlsx_path)
    wb = storage._open_or_create()
    ws_trade = wb["trade_records"]
    ws_raw = wb["raw_data"]

    storage._append_raw(ws_raw, collect_time, json_path, raw_text)
    storage._autosize(wb)
    wb.save(storage.xlsx_path)

    stats = {
        "trade_records_added": 0,
        "trade_records_total": max(ws_trade.max_row - 1, 0),
        "raw_data_rows_added": 1,
        "raw_data_rows_total": max(ws_raw.max_row - 1, 0),
    }
    logger.info(
        f"[v2] 持久化完成 → {storage.xlsx_path}："
        f"trade_records 未追加（留给 AIParser 阶段），累计 {stats['trade_records_total']} 行，"
        f"raw_data +{stats['raw_data_rows_added']} 行，累计 {stats['raw_data_rows_total']} 行。"
    )
    return stats


# ============================================================
#  独立 CLI
# ============================================================

def _cli():
    ap = argparse.ArgumentParser(description="独立 CLI：把采集 JSON 追加到 Excel 数据库（每次一行 raw_data）")
    ap.add_argument("json", nargs="?", help="采集生成的 JSON 路径")
    ap.add_argument("--json", dest="json_opt", metavar="PATH", help="命名参数形式：采集生成的 JSON 路径")
    ap.add_argument("--xlsx", metavar="PATH", default=None, help=f"目标 Excel（默认: output/{DEFAULT_XLSX_BASENAME}）")
    ap.add_argument("--data-source", default=DEFAULT_DATA_SOURCE, help=f"数据来源，默认 {DEFAULT_DATA_SOURCE!r}")
    args = ap.parse_args()

    json_path = args.json or args.json_opt
    if not json_path:
        ap.print_help()
        sys.exit(2)
    json_path = os.path.abspath(json_path)
    if not os.path.exists(json_path):
        print(f"[ERROR] JSON 不存在: {json_path}")
        sys.exit(3)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-7s | %(name)s | %(message)s",
        datefmt="%H:%M:%S",
    )

    stats = persist_json_to_excel(json_path, xlsx_path=args.xlsx, data_source=args.data_source)
    print("\n✅ 持久化完成：")
    print(f"   JSON 结果 : {json_path}")
    xlsx = os.path.abspath(args.xlsx or os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "output",
        DEFAULT_XLSX_BASENAME,
    ))
    print(f"   Excel 更新: {xlsx}")
    print(f"   新增记录  : {stats['trade_records_added']} 条 (总 {stats['trade_records_total']})")
    print(f"   raw_data  : {stats['raw_data_rows_total']} 条")


if __name__ == "__main__":
    _cli()
