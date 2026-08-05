"""
蚂蚁财富大V交易数据 → Excel 导出模块

功能：
  1. 接受 采集生成的 JSON 文件路径 或 List[dict]（records / collected_trades 列表）
  2. 输出到 output/ant_finance_trades.xlsx（4 个 Sheet）
  3. 追加写入，不覆盖历史数据
  4. 按 {大V名称 + 时间 + 操作类型 + 基金 + 金额} 去重
  5. Excel 不存在时自动创建

使用方式（直接独立跑）：
  python core/excel_exporter.py output/kol_trades_20260804_235016.json
  python core/excel_exporter.py --json output/kol_trades_XXX.json --xlsx output/ant_finance_trades.xlsx
"""
from __future__ import annotations

import os
import sys
import json
import argparse
import datetime
import logging
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


logger = logging.getLogger("excel_exporter")


# ============================================================
#  常量：各 Sheet 的字段定义
# ============================================================

TRADE_FIELDS: List[str] = [
    "data_source",        # 新增：数据来源（默认"蚂蚁财富"）
    "采集日期",
    "大V名称",
    "操作时间",
    "操作类型",
    "基金名称",
    "源基金",
    "目标基金",
    "操作金额",
    "金额单位",
    "观点文本",
    "点赞数",
    "评论数",
    "置信度",
    "采集时间",
]

KOL_INFO_FIELDS: List[str] = [
    "大V名称",
    "是否启用",
    "采集优先级",
    "备注",
]

DAILY_SUMMARY_FIELDS: List[str] = [
    "日期",
    "方向",
    "操作人数",
    "操作次数",
    "累计金额",
]

RAW_DATA_FIELDS: List[str] = [
    "采集时间",
    "大V名称",
    "operation_text",
    "原始JSON",
]


DEFAULT_DATA_SOURCE = "蚂蚁财富"

# Excel 默认文件名（相对项目根的 output 目录）
DEFAULT_XLSX_FILENAME = "ant_finance_trades.xlsx"


# ============================================================
#  辅助函数
# ============================================================

def _safe_str(v: Any, default: str = "") -> str:
    if v is None:
        return default
    if isinstance(v, float):
        # 整数值不带小数点
        if v.is_integer():
            return str(int(v))
        return f"{v:.4f}".rstrip("0").rstrip(".")
    return str(v)


def _safe_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace(",", "").strip()
        return float(s)
    except (ValueError, TypeError):
        return default


def _strip_amount_text(text: str) -> str:
    """
    去掉金额字符串里的非数字部分：
    '加仓2000元' → '2000'
    '近一年收益率31.77%' → '31.77'（如果最终误识别也保留纯数字）
    '5,008.00' → '5008.00'
    """
    import re
    if not text:
        return ""
    m = re.search(r"([\d,]+(?:\.\d+)?)", text.replace("，", ","))
    return m.group(1).replace(",", "") if m else text.replace(",", "")


def _pick_amount_value(rec: Dict[str, Any]) -> str:
    """从 record 中获取干净的金额数值字符串"""
    # 优先用解析过的数值字段（KolOperation → to_dict 输出）
    av = rec.get("amount_value") if rec.get("amount_value") not in (None, "", 0) else None
    # TRANSFER 时还有 *_amount_value
    if av is None:
        av = rec.get("source_amount_value")
    if av is None:
        av = rec.get("target_amount_value")
    if av is not None and av not in ("", 0):
        return _safe_str(av)
    # 回退：解析原始 amount/source_amount/target_amount 文本
    for k in ("amount", "source_amount", "target_amount"):
        txt = rec.get(k)
        if txt:
            v = _strip_amount_text(txt)
            if v:
                return v
    return ""


def _pick_unit(rec: Dict[str, Any]) -> str:
    at = rec.get("action_type", "")
    u = rec.get("unit") or rec.get("amount_unit")
    if u:
        return u
    if at == "BUY":
        src_u = rec.get("target_amount_unit") or rec.get("target_amount")
        if rec.get("target_amount_unit"):
            return rec["target_amount_unit"]
        return "元"
    if at == "SELL":
        src_u = rec.get("source_amount_unit")
        if src_u:
            return src_u
        return "份"
    if at == "TRANSFER":
        # TRANSFER 一行存两个金额；这里取"份"（源），单位列做展示，详细看源/目标
        return rec.get("source_amount_unit") or "份/元"
    if at == "CANCEL":
        return rec.get("unit") or rec.get("amount_unit") or ""
    return ""


def _pick_fund(rec: Dict[str, Any]) -> str:
    at = rec.get("action_type", "")
    if at == "TRANSFER":
        return ""  # TRANSFER 用源基金/目标基金列
    if at == "CANCEL":
        return _safe_str(rec.get("cancel_type")) + (("  " + _safe_str(rec.get("fund"))) if rec.get("fund") else "")
    return _safe_str(rec.get("fund") or rec.get("fund_name"))


def _dedup_key(rec: Dict[str, Any]) -> Tuple[str, str, str, str, str]:
    """
    去重键：大V名称 + 时间 + 操作类型 + 基金 + 金额
    各字段做标准化（去空格、小写金额千分位等）
    """
    kol = _safe_str(rec.get("kol_name")).strip()
    ts = _safe_str(rec.get("timestamp")).strip()
    at = _safe_str(rec.get("action_type") or rec.get("operation")).strip().upper()

    # 基金：BUY/SELL 取 fund，TRANSFER 取 source_fund + target_fund，CANCEL 取 cancel_type+fund
    if at == "TRANSFER":
        fund = f"{_safe_str(rec.get('source_fund'))}|{_safe_str(rec.get('target_fund'))}"
    elif at == "CANCEL":
        fund = f"{_safe_str(rec.get('cancel_type'))}|{_safe_str(rec.get('fund') or rec.get('fund_name'))}"
    else:
        fund = _safe_str(rec.get("fund") or rec.get("fund_name"))

    amt = _pick_amount_value(rec)
    return (kol, ts, at, fund.strip(), amt.strip())


# ============================================================
#  核心：ExcelExporter
# ============================================================

class ExcelExporter:
    """
    把采集结果追加写入 Excel（4 个 Sheet）。

    典型用法：
        ex = ExcelExporter("output/ant_finance_trades.xlsx")
        ex.export_json("output/kol_trades_20260804_235016.json")
        # 或者：
        ex.export_records([{...}, {...}], generated_at="2026-08-04T23:50:16")
    """

    def __init__(self, xlsx_path: str):
        self.xlsx_path = os.path.abspath(xlsx_path)
        self._ensure_dir()

    # ----------------------------------------------------------
    def _ensure_dir(self) -> None:
        d = os.path.dirname(self.xlsx_path)
        if d:
            os.makedirs(d, exist_ok=True)

    # ----------------------------------------------------------
    #  输入：JSON → list[dict]
    # ----------------------------------------------------------
    def load_json_records(self, json_path: str) -> Tuple[List[Dict[str, Any]], str]:
        """
        读取采集生成的 JSON。
        兼容两种根格式：
          A. main.py argparse 输出：{"records": [...], "generated_at": ...}
          B. ScrollManager 输出：{"collected_trades": [...], "generated_at": ...}

        Returns: (records, generated_at_str)
        """
        with open(json_path, "r", encoding="utf-8") as f:
            payload = json.load(f)

        recs: List[Dict[str, Any]] = []
        if isinstance(payload, list):
            recs = [r for r in payload if isinstance(r, dict)]
            gen_at = ""
        elif isinstance(payload, dict):
            recs = (
                payload.get("records")
                or payload.get("collected_trades")
                or payload.get("operations")
                or []
            )
            gen_at = _safe_str(payload.get("generated_at") or payload.get("collected_at") or "")
        else:
            raise ValueError(f"未知 JSON 结构: {type(payload)}")
        return recs, gen_at

    # ----------------------------------------------------------
    #  对外：主入口
    # ----------------------------------------------------------
    def export_json(self, json_path: str, data_source: str = DEFAULT_DATA_SOURCE) -> Dict[str, int]:
        """从 JSON 文件导出，返回统计 {sheet名: 新增行数}"""
        recs, gen_at = self.load_json_records(json_path)
        return self.export_records(recs, generated_at=gen_at, data_source=data_source)

    def export_records(
        self,
        records: Iterable[Dict[str, Any]],
        generated_at: str = "",
        data_source: str = DEFAULT_DATA_SOURCE,
    ) -> Dict[str, int]:
        records = list(records)
        if not generated_at:
            generated_at = datetime.datetime.now().isoformat(timespec="seconds")

        collect_date, collect_time = self._split_datetime(generated_at)

        wb = self._open_or_create_workbook()
        ws_trade = wb["trade_records"]
        ws_kol = wb["kol_info"]
        ws_daily = wb["daily_summary"]
        ws_raw = wb["raw_data"]

        # 1) 读取 trade_records 中已有去重键，避免重复
        existing_keys = self._load_existing_dedup_keys(ws_trade)

        # 2) 写入 trade_records & raw_data & 收集 kol 列表
        trade_added = 0
        raw_added = 0
        kol_set = set()

        for rec in records:
            if not isinstance(rec, dict):
                continue
            key = _dedup_key(rec)
            if key in existing_keys:
                continue
            existing_keys.add(key)

            trade_row = self._build_trade_row(rec, collect_date, collect_time, data_source)
            ws_trade.append(trade_row)
            trade_added += 1

            raw_row = self._build_raw_row(rec, collect_time, data_source, orig_dict=rec)
            ws_raw.append(raw_row)
            raw_added += 1

            kol_name = _safe_str(rec.get("kol_name")).strip()
            if kol_name:
                kol_set.add(kol_name)

        # 3) kol_info：补齐未出现过的大V（保持不覆盖）
        kol_added = self._append_new_kols(ws_kol, kol_set)

        # 4) daily_summary：预留（不做自动统计，避免口径误判；至少保留表头；本次不自动追加）
        #    用户可后续用 AI / 公式在该 Sheet 上做透视。
        daily_added = 0

        self._autosize_and_style(wb)
        wb.save(self.xlsx_path)

        stats = {
            "trade_records": trade_added,
            "kol_info": kol_added,
            "daily_summary": daily_added,
            "raw_data": raw_added,
        }
        logger.info(
            f"Excel 追加完成 -> {self.xlsx_path}："
            + ", ".join(f"{k} +{v}" for k, v in stats.items())
        )
        return stats

    # ----------------------------------------------------------
    #  内部：Workbook / Sheet 管理
    # ----------------------------------------------------------
    def _open_or_create_workbook(self) -> Workbook:
        if os.path.exists(self.xlsx_path):
            wb = load_workbook(self.xlsx_path)
            # 确保 4 个 Sheet 都存在
            for name, fields in [
                ("trade_records", TRADE_FIELDS),
                ("kol_info", KOL_INFO_FIELDS),
                ("daily_summary", DAILY_SUMMARY_FIELDS),
                ("raw_data", RAW_DATA_FIELDS),
            ]:
                if name not in wb.sheetnames:
                    ws = wb.create_sheet(title=name)
                    self._write_header(ws, fields)
            return wb
        else:
            wb = Workbook()
            # 默认 Sheet 改名为 trade_records
            ws0 = wb.active
            ws0.title = "trade_records"
            self._write_header(ws0, TRADE_FIELDS)

            ws_kol = wb.create_sheet("kol_info")
            self._write_header(ws_kol, KOL_INFO_FIELDS)

            ws_daily = wb.create_sheet("daily_summary")
            self._write_header(ws_daily, DAILY_SUMMARY_FIELDS)

            ws_raw = wb.create_sheet("raw_data")
            self._write_header(ws_raw, RAW_DATA_FIELDS)
            return wb

    @staticmethod
    def _write_header(ws, fields: List[str]) -> None:
        # 先判断当前第 1 行是否真的有内容（任意 cell.value 非 None 即视为有表头）
        already_has_header = any(
            ws.cell(row=1, column=i).value is not None
            for i in range(1, len(fields) + 1)
        )
        if not already_has_header:
            for i, f in enumerate(fields, 1):
                ws.cell(row=1, column=i, value=f)
            # 表头加粗 + 底色
            bold = Font(bold=True, color="FFFFFF")
            fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center")
            for col in range(1, len(fields) + 1):
                c = ws.cell(row=1, column=col)
                c.font = bold
                c.fill = fill
                c.alignment = center
            ws.freeze_panes = "A2"
        # 如果已有表头但列不一致（未来新增列），不做破坏性改动，避免覆盖历史

    # ----------------------------------------------------------
    def _load_existing_dedup_keys(self, ws) -> set:
        keys: set = set()
        if ws.max_row <= 1:
            return keys
        # 读取表头，找到对应列号（列顺序保持与 TRADE_FIELDS 一致即可，不必按表头文字索引）
        header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        # 建立名字 → 索引
        idx = {str(h): i for i, h in enumerate(header, 1) if h is not None}

        def col(name: str) -> Optional[int]:
            # 中文列名优先；找不到再按 TRADE_FIELDS 相对位置兜底
            if name in idx:
                return idx[name]
            if name in TRADE_FIELDS:
                pos = TRADE_FIELDS.index(name) + 1
                if pos <= len(header):
                    return pos
            return None

        for r in range(2, ws.max_row + 1):
            def gv(name):
                c = col(name)
                return _safe_str(ws.cell(row=r, column=c).value) if c else ""

            kol = gv("大V名称").strip()
            ts = gv("操作时间").strip()
            at = gv("操作类型").strip().upper()
            fund = ""
            if at == "TRANSFER":
                fund = f"{gv('源基金').strip()}|{gv('目标基金').strip()}"
            elif at.startswith("BUY_CANCEL") or at.startswith("SELL_CANCEL") or at.startswith("TRANSFER_CANCEL") or (
                "CANCEL" in at
            ):
                # CANCEL 在写入时把 cancel_type 写在基金列的开头，所以直接读基金列
                fund = gv("基金名称").strip()
            else:
                fund = gv("基金名称").strip()
            amt = _strip_amount_text(gv("操作金额")).strip()
            keys.add((kol, ts, at, fund, amt))
        return keys

    # ----------------------------------------------------------
    #  行构造
    # ----------------------------------------------------------
    @staticmethod
    def _split_datetime(iso_or_dt: str) -> Tuple[str, str]:
        s = (iso_or_dt or "").strip().replace("T", " ")
        # "2026-08-04T23:50:16" / "2026-08-04 23:50:16"
        try:
            if len(s) >= 19:
                dt = datetime.datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
            elif len(s) >= 10:
                dt = datetime.datetime.strptime(s[:10], "%Y-%m-%d")
            else:
                raise ValueError
        except Exception:
            dt = datetime.datetime.now()
        return dt.strftime("%Y-%m-%d"), dt.strftime("%Y-%m-%d %H:%M:%S")

    def _build_trade_row(
        self, rec: Dict[str, Any], collect_date: str, collect_time: str, data_source: str
    ) -> List[Any]:
        kol = _safe_str(rec.get("kol_name")).strip()
        # 操作时间：**只在 record 自身 timestamp 非空时才填**，否则留空；
        # 这样去重时和 JSON 侧构造的 key 完全对称（空 ↔ 空），避免因填充采集时间导致去重失效。
        ts = _safe_str(rec.get("timestamp")).strip()
        at = _safe_str(rec.get("action_type") or rec.get("operation")).strip().upper()

        # 基金
        fund = _pick_fund(rec)
        source_fund = _safe_str(rec.get("source_fund"))
        target_fund = _safe_str(rec.get("target_fund"))
        if at == "TRANSFER":
            # TRANSFER 独立用源基金/目标基金列，基金名称列留空
            fund = ""
        if at == "CANCEL" and fund:
            # CANCEL：把 cancel_type 与 fund 用 "|" 合并，
            # 这样去重时 Excel→key 与 JSON→key 完全对称
            cancel_t = _safe_str(rec.get("cancel_type") or "UNKNOWN")
            raw_fund = _safe_str(rec.get("fund") or rec.get("fund_name"))
            fund = f"{cancel_t}|{raw_fund}"

        # 金额：BUY/SELL/CANCEL 取单笔；TRANSFER 取"源份额/目标金额"拼接
        if at == "TRANSFER":
            src = _pick_amount_value({"amount_value": rec.get("source_amount_value"), "amount": rec.get("source_amount")})
            tgt = _pick_amount_value({"amount_value": rec.get("target_amount_value"), "amount": rec.get("target_amount")})
            if src and tgt:
                amount_str = f"{src} / {tgt}"
            else:
                amount_str = src or tgt or ""
            unit_str = (
                f"{rec.get('source_amount_unit') or '份'} / {rec.get('target_amount_unit') or '元'}"
            )
        else:
            amount_str = _pick_amount_value(rec)
            unit_str = _pick_unit(rec)

        # 观点文本（从 operation_text 中剥离掉纯交易锚点+金额行，保留可能在同一节点的观点；
        #  若 operation_text 纯交易文本（短），就直接存；太长时截断）
        opinion = _safe_str(rec.get("operation_text"))
        if len(opinion) > 500:
            opinion = opinion[:500] + "…[已截断，完整见raw_data]"

        # 点赞/评论：当前采集不到，留空（后续可扩展）
        like_count = ""
        comment_count = ""
        conf = _safe_float(rec.get("confidence"))

        return [
            data_source,           # data_source
            collect_date,          # 采集日期
            kol,                   # 大V名称
            ts,                    # 操作时间
            at,                    # 操作类型
            fund,                  # 基金名称
            source_fund,           # 源基金
            target_fund,           # 目标基金
            amount_str,            # 操作金额
            unit_str,              # 金额单位
            opinion,               # 观点文本
            like_count,            # 点赞数
            comment_count,         # 评论数
            conf,                  # 置信度
            collect_time,          # 采集时间
        ]

    def _build_raw_row(
        self, rec: Dict[str, Any], collect_time: str, data_source: str, orig_dict: Dict[str, Any]
    ) -> List[Any]:
        kol = _safe_str(rec.get("kol_name")).strip()
        op_text = _safe_str(rec.get("operation_text"))
        # 原始 JSON：用紧凑形式，截断超大文本（>32k 的单元格 Excel 易报错）
        try:
            raw_json = json.dumps(orig_dict, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            raw_json = str(orig_dict)
        MAX = 32000
        if len(raw_json) > MAX:
            raw_json = raw_json[: MAX - 12] + "…[TRUNCATED]"
        return [
            collect_time,
            kol,
            op_text,
            raw_json,
        ]

    def _append_new_kols(self, ws, kol_set) -> int:
        # 读取已有大V列（第1列），避免重复写入
        existing = set()
        if ws.max_row >= 2:
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if v:
                    existing.add(str(v).strip())
        # 进一步剔除 settings 里的 KOL 黑名单（如果存在）
        try:
            from config import settings as _settings
            kol_blacklist = set()
            for _name in ("INVALID_KOL_NAMES", "INVALID_FUND_NAMES"):
                arr = getattr(_settings, _name, ()) or ()
                if _name == "INVALID_KOL_NAMES":
                    kol_blacklist.update(str(x).strip() for x in arr if x)
        except Exception:
            kol_blacklist = set()
        added = 0
        for kol in sorted(kol_set):
            if kol in existing:
                continue
            if kol in kol_blacklist:
                # 黑名单 KOL（如"全部"、"最新"）不入库
                continue
            # 是否启用=是，优先级=5（默认中），备注=自动新增
            ws.append([kol, "是", 5, "系统自动导入"])
            existing.add(kol)
            added += 1
        return added

    # ----------------------------------------------------------
    #  样式：自动列宽 + 金额/置信度单元格格式
    # ----------------------------------------------------------
    def _autosize_and_style(self, wb: Workbook) -> None:
        # 自动列宽（仅对各 Sheet 前 ~200 行做估算）
        widths_map = {
            "trade_records": [10, 12, 16, 20, 10, 34, 34, 34, 14, 10, 60, 8, 8, 8, 20],
            "kol_info": [20, 8, 10, 24],
            "daily_summary": [12, 10, 10, 10, 14],
            "raw_data": [20, 16, 60, 80],
        }
        for sheet_name, widths in widths_map.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # 长文本单元格自动换行（主要是观点文本和 raw_data）
            if sheet_name == "trade_records":
                op_col = TRADE_FIELDS.index("观点文本") + 1
                wrap = Alignment(wrap_text=True, vertical="top")
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=op_col).alignment = wrap
                # 置信度用 0.000 显示
                conf_col = TRADE_FIELDS.index("置信度") + 1
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=conf_col).number_format = "0.000"
            if sheet_name == "raw_data":
                ot_col = RAW_DATA_FIELDS.index("operation_text") + 1
                json_col = RAW_DATA_FIELDS.index("原始JSON") + 1
                wrap = Alignment(wrap_text=True, vertical="top")
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=ot_col).alignment = wrap
                    ws.cell(row=r, column=json_col).alignment = wrap


# ============================================================
#  CLI：直接 python core/excel_exporter.py xxx.json
# ============================================================

def _cli():
    ap = argparse.ArgumentParser(description="独立 CLI：把 JSON 采集结果追加到 Excel")
    ap.add_argument("json", nargs="?", help="采集生成的 JSON 路径")
    ap.add_argument("--json", dest="json_opt", metavar="PATH", help="采集生成的 JSON 路径（命名参数形式）")
    ap.add_argument("--xlsx", metavar="PATH", default=None, help=f"目标 Excel 路径（默认: output/{DEFAULT_XLSX_FILENAME}）")
    ap.add_argument("--data-source", default=DEFAULT_DATA_SOURCE, help=f"data_source 字段值（默认: {DEFAULT_DATA_SOURCE}）")
    args = ap.parse_args()

    json_path = args.json or args.json_opt
    if not json_path:
        ap.print_help()
        sys.exit(2)
    if not os.path.isabs(json_path):
        json_path = os.path.abspath(json_path)
    if not os.path.exists(json_path):
        print(f"[ERROR] JSON 不存在: {json_path}")
        sys.exit(3)

    if args.xlsx:
        xlsx_path = args.xlsx
    else:
        # 默认：项目根下的 output/ant_finance_trades.xlsx
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xlsx_path = os.path.join(project_root, "output", DEFAULT_XLSX_FILENAME)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-7s | %(name)s | %(message)s",
        datefmt="%H:%M:%S",
    )

    ex = ExcelExporter(xlsx_path)
    stats = ex.export_json(json_path, data_source=args.data_source)
    print("✅ 导出完成：")
    print(f"   JSON  : {json_path}")
    print(f"   Excel : {xlsx_path}")
    for k, v in stats.items():
        print(f"   {k:<14s}: +{v} 行")


if __name__ == "__main__":
    _cli()
